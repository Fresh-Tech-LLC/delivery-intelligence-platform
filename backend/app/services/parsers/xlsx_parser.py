"""Microsoft Excel (.xlsx) parser."""
from __future__ import annotations

from pathlib import Path
from typing import Any


def extract_text(path: Path) -> tuple[str, dict[str, Any]]:
    """Return (extracted_text, metadata_dict) for .xlsx files.

    Uses openpyxl in read_only mode for speed. wb.close() is mandatory in
    read_only mode — zip file handles remain open otherwise.
    None cell values are converted to empty string to avoid "None" literals.
    """
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    wb = load_workbook(str(path), data_only=True, read_only=True)
    parts: list[str] = []
    sheet_names: list[str] = []
    try:
        for sheet in wb.worksheets:
            sheet_names.append(sheet.title)  # collect here — wb.worksheets inaccessible after close()
            parts.append(f"[Sheet: {sheet.title}]")
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join("" if v is None else str(v) for v in row)
                if row_text.strip():
                    parts.append(row_text)
    finally:
        wb.close()
    return "\n".join(parts), {
        "sheet_names": sheet_names,
        "sheet_count": len(sheet_names),
        "parser": "xlsx_parser",
    }
